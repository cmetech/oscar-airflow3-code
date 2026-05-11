import logging
import base64
import json
import openpyxl
import pandas as pd
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, Optional, List
from airflow import DAG
import pendulum
from airflow.operators.python import PythonOperator
from helpers.schema.email_processing import AttachmentStorageType  # type: ignore
from helpers.email_helper import retrieve_stored_attachment  # type: ignore
from hooks.worklog_hook import WorkLogHook, WorkLogType  # type: ignore
from hooks.rule_hook import RuleHook  # type: ignore
from hooks.alert_hook import AlertHook  # type: ignore
import re
from io import BytesIO
from hooks.cache_hook import CacheHook  # type: ignore
import asyncio
import os
import uuid
import httpx
from hooks.access_management_db_hook import AccessManagementSQLHook
from jinja2 import Template
from hooks.mapping_hook import MappingHook
from hooks.notify_hook import NotifyHook

logger = logging.getLogger(__name__)

default_args = {
    "owner": "airflow",
    "depends_on_past": False,
    "email_on_failure": False,
    "email_on_retry": False,
    "retries": 1,
    "retry_delay": timedelta(minutes=5),
}

# Reuse the same EXCEL_TYPE_CONFIGS from excel_to_db_ingest_dag.py
EXCEL_TYPE_CONFIGS = {
    'platform_role_mapping': {
        'filename': 'Platform_Role_Mapping.xlsx',
        'sheet': 'Sheet1',
        'table': 'EXT_AM_ITSM_ROLE_PLATFORM_MAPPING',
        'mapping': {
            'Role Description': 'u_role_description',
            'Company': 'u_company',
            'Platform': 'u_platform',
            'Role Display Name': 'u_role',
        },
    },
    'approver_role': {
        'filename': 'Approver_Roles.xlsx',
        'sheet': 'Report',
        'table': 'EXT_AM_ITSM_APPROVER_ROLE',
        'mapping': {
            'Company': 'u_company',
            'Platform': 'u_platform',
            'Role Description': 'u_role_des',
            'Approver List': 'u_approver_list',
            'Approver Names': 'u_approver_name',
            'Display Name': 'u_display_name',
        },
    },
    'user_report': {
        'filename': 'UserReportITSM_ENV.xlsx',
        'sheet': 'Report',
        'table': 'EXT_AM_ITSM_USER_LIST',
        'mapping': {
            'Login ID': 'u_login_id',
            'First Name*': 'u_first_name',
            'Last Name*+': 'u_last_name',
            'Full Name': 'u_full_name',
            'Email Address': 'u_email',
            'Environment': 'u_env',
        },
    },
    'shift_roster': {
        'filename': 'SampleShiftRoster.xlsx',
        'sheet': 'Sheet1',
        'table': 'EXT_AM_SHIFT_ROSTER_TABLE',
        'mapping': {
            'Name': 'u_name',
            'Email': 'u_email',
            'Date': 'u_current_date',
            'Shift Time': 'u_full_time',
        },
    },
}

# Connection ID for DB (set via Airflow connection/environment)
DB_CONN_ID = os.getenv('OSCAR_DB_EXT_CONNECTION_ID', 'oscar_db_ext')
ENABLE_AM_SUPPORT_DATA_PROCESSING = os.getenv('ENABLE_AM_SUPPORT_DATA_PROCESSING', 'True').lower() == 'true'

# Mapping from internal field names to user-friendly names for emails
FIELD_NAME_MAP = {
    "user_adid": "User's AD ID",
    "nt_id": "User's NT ID",
    "email": "Email Address",
    "organization": "Organization",
    "description": "Description",
    "owner_group_email": "Owner Group Email",
    "name": "Name",
    "user_full_name": "User's Full Name",
    "service_account_full_name": "Service Account Full Name",
    "service_account_owner_name": "Service Account Owner Name",
    "manager_name": "Manager Name",
    "team_name": "Team Name",
    "user_email_id": "User Email ID",
    "service_account_owner_email_id": "Service Account Owner Email ID",
    "manager_email_id": "Manager's Email ID",
    "user_ntid_signum": "User's NTID Signum",
    "user_ntid": "User's NTID",
    "owner_ntid": "Owner's NTID",
    "user_uprising_id": "User's Uprising ID",
    "user_ad_id": "User's AD ID",
    "manager_ad_id": "Manager's AD ID"
    # Add more mappings as needed
}

def determine_excel_type(email_subject):
    """Determine excel_type based on email subject content."""
    email_subject = email_subject.lower()
    if 'platform role mapping' in email_subject:
        return 'platform_role_mapping'
    elif 'approver role' in email_subject:
        return 'approver_role'
    elif 'user report' in email_subject:
        return 'user_report'
    elif 'shift roster' in email_subject:
        return 'shift_roster'
    else:
        raise ValueError(f"Could not determine excel_type from email subject: {email_subject}")

def save_mapping_excel_data_to_db(excel_content_base64, email_subject, worklog_id):
    """
    Process Excel data from base64 content and save to appropriate database table.

    Args:
        excel_content_base64 (str): Base64 encoded Excel file content
        email_subject (str): Email subject used to determine excel_type

    Returns:
        dict: Processing results including row count and status
    """
    db_hook = None
    connection = None
    try:
        # Determine excel_type from email subject
        excel_type = determine_excel_type(email_subject)
        config = EXCEL_TYPE_CONFIGS.get(excel_type)
        if not config:
            msg = f"Unknown excel_type '{excel_type}'. Valid types: {list(EXCEL_TYPE_CONFIGS.keys())}"
            logger.error(msg)
            return {"status": "error", "message": msg}

        # Process base64 content
        logger.info(f"Reading Excel for type '{excel_type}' from base64 content")
        excel_bytes = base64.b64decode(excel_content_base64)
        excel_file = BytesIO(excel_bytes)
        df = pd.read_excel(excel_file, sheet_name=config['sheet'])

        # Process and transform data
        df = df.rename(columns=config['mapping'])
        df = df[list(config['mapping'].values())]
        df = df.where(pd.notnull(df), None)

        # Database operations with transaction management
        db_hook = AccessManagementSQLHook(DB_CONN_ID, worklog_id)
        table = config['table']

        # Get database connection and start transaction
        connection = db_hook.hook.get_conn()
        cursor = connection.cursor()

        # Disable autocommit to start transaction
        connection.autocommit = False
        logger.info(f"Started transaction for table {table}")

        # Create table if not exists
        columns_defs = ', '.join([
            f"{col} DATE" if 'date' in col.lower() else f"{col} VARCHAR(255)" for col in df.columns
        ])
        create_table_sql = f"CREATE TABLE IF NOT EXISTS {table} ({columns_defs})"
        cursor.execute(create_table_sql)
        logger.info(f"Ensured table {table} exists with columns: {', '.join(df.columns)}")

        # Create temporary table for new data
        temp_table = f"{table}_temp_{int(datetime.now().timestamp())}"
        create_temp_table_sql = f"CREATE TABLE {temp_table} LIKE {table}"
        cursor.execute(create_temp_table_sql)
        logger.info(f"Created temporary table {temp_table}")

        # Insert new data into temporary table
        data_tuples = [tuple(x) for x in df.values]
        placeholders = ', '.join(['%s'] * len(df.columns))
        insert_sql = f"INSERT INTO {temp_table} ({', '.join(df.columns)}) VALUES ({placeholders})"

        cursor.executemany(insert_sql, data_tuples)
        logger.info(f"Successfully inserted {len(data_tuples)} rows into temporary table {temp_table}")

        # Verify data integrity in temporary table
        cursor.execute(f"SELECT COUNT(*) FROM {temp_table}")
        result = cursor.fetchone()
        if result[0] != len(df):
            raise ValueError(f"Data integrity check failed: expected {len(df)} rows, got {result[0]}")

        # Replace old table with new data using atomic operation
        # First, backup existing data by renaming current table
        backup_table = f"{table}_backup_{int(datetime.now().timestamp())}"
        cursor.execute(f"RENAME TABLE {table} TO {backup_table}")
        logger.info(f"Backed up existing data to {backup_table}")

        # Rename temporary table to main table
        cursor.execute(f"RENAME TABLE {temp_table} TO {table}")
        logger.info(f"Renamed temporary table to main table {table}")

        # Commit transaction
        connection.commit()
        logger.info(f"Transaction committed successfully for {table}")

        # Drop backup table after successful commit
        try:
            cursor.execute(f"DROP TABLE {backup_table}")
            logger.info(f"Dropped backup table {backup_table}")
        except Exception as e:
            logger.warning(f"Failed to drop backup table {backup_table}: {str(e)}")

        return {
            'status': 'success',
            'excel_type': excel_type,
            'rows_processed': len(df),
            'table': table
        }

    except Exception as e:
        # Rollback transaction on any error
        if connection:
            try:
                connection.rollback()
                logger.info("Transaction rolled back due to error")
            except Exception as rollback_error:
                logger.error(f"Error during rollback: {str(rollback_error)}")

        error_msg = f"Error processing Excel data: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {
            'status': 'error',
            'message': error_msg
        }
    finally:
        # Clean up resources
        if connection:
            try:
                connection.autocommit = True
                connection.close()
            except Exception as e:
                logger.warning(f"Error closing connection: {str(e)}")

# def create_alert_object(request_data: Dict[str, Any], alert_type: Optional[str] = None) -> Dict[str, Any]:
#     pass


async def process_attachment(
    attachment: Dict[str, Any], request_data: Dict[str, Any], worklog_hook: WorkLogHook
) -> Dict[str, Any]:
    """
    Process an attachment based on its storage type and content type.

    Args:
        attachment: The attachment dictionary containing storage type, content, etc.
        request_data: The complete request data containing email information
        worklog_hook: The worklog hook to log processing information

    Returns:
        Dict containing the processed attachment data
    """
    filename = attachment.get("filename", "unknown")
    content_type = attachment.get("content_type", "")
    storage_type = attachment.get("storage_type", AttachmentStorageType.INLINE)
    size = attachment.get("size", 0)

    # Short-circuit: attachment already failed during storage (e.g. Redis store error in taskmanager)
    if attachment.get("status") == "error":
        err_msg = attachment.get("error") or "Attachment storage failed before retrieval"
        logger.error(f"Attachment {filename!r} arrived with pre-existing error status: {err_msg!r}")
        worklog_hook.error(f"Attachment {filename} could not be retrieved — storage error: {err_msg}")
        return {
            "filename": filename,
            "content_type": content_type,
            "status": "error",
            "error": err_msg,
        }

    logger.info(f"Processing attachment: {filename} ({content_type}, {size} bytes)")

    worklog_hook.info(f"Processing attachment: {filename} ({content_type}, {size} bytes)")

    # Get the attachment content based on storage type
    attachment_content = None

    if storage_type == AttachmentStorageType.INLINE:
        # For inline attachments, the content is already base64 encoded
        logger.info(f"Storage type {storage_type}")
        attachment_content = attachment.get("content")
        if not attachment_content:
            worklog_hook.warning(f"No content found for inline attachment: {filename}")
            logger.info(f"No content found for inline attachment: {filename}")
            return {
                "filename": filename,
                "content_type": content_type,
                "status": "error",
                "error": "No content found for inline attachment",
            }
    elif storage_type == AttachmentStorageType.REDIS:
        # For Redis stored attachments, we need to retrieve them
        logger.info(f"Storage type {storage_type}")
        reference_id = attachment.get("reference_id")
        if not reference_id:
            logger.info(f"No reference_id found for Redis attachment: {filename} content-type: {content_type}")
            worklog_hook.warning(f"No reference_id found for Redis attachment: {filename}")
            return {
                "filename": filename,
                "content_type": content_type,
                "status": "error",
                "error": "No reference_id found for Redis attachment",
            }

        try:
            async with CacheHook() as cache_hook:
                binary_data = await cache_hook.get_binary_content(reference_id)
                if binary_data and "data" in binary_data:
                    attachment_content = binary_data["data"]
                    # Get original content type from metadata if available
                    if "metadata" in binary_data and "original_content_type" in binary_data["metadata"]:
                        content_type = binary_data["metadata"]["original_content_type"]
                        worklog_hook.info(f"Retrieved original content type: {content_type}")
                        logger.info(f"Retrieved original content type: {content_type}")
                    worklog_hook.info(f"Successfully retrieved attachment from Redis: {filename}")
                    logger.info(f"Successfully retrieved attachment from Redis: {filename}")
                else:
                    worklog_hook.warning(f"Failed to retrieve content from Redis for reference_id: {reference_id}")
                    logger.info(f"Failed to retrieve content from Redis for reference_id: {reference_id}")
                    return {
                        "filename": filename,
                        "content_type": content_type,
                        "status": "error",
                        "error": "Failed to retrieve content from Redis",
                    }
        except Exception as e:
            worklog_hook.error(f"Error retrieving attachment from Redis: {str(e)}")
            logger.error(f"Error retrieving attachment from Redis: {str(e)}")
            return {
                "filename": filename,
                "content_type": content_type,
                "status": "error",
                "error": f"Error retrieving attachment from Redis: {str(e)}",
            }
    else:
        worklog_hook.warning(f"Unsupported storage type: {storage_type}")
        logger.info(f"Unsupported storage type: {storage_type}")
        return {
            "filename": filename,
            "content_type": content_type,
            "status": "error",
            "error": f"Unsupported storage type: {storage_type}",
        }

    # Process the attachment based on content type
    logger.info(f"Content type is {content_type}")
    if (
        content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or content_type == "application/vnd.ms-excel"
    ):
        # This is an Excel file, process it
        worklog_hook.info(f"Processing Excel file: {filename}")
        logger.info(f"Processing Excel file: {filename}")
        try:
            # Decode the base64 content
            decoded_content = base64.b64decode(attachment_content)

            # Get email subject and extract effective subject key
            subject = request_data.get("subject", "").lower()
            # Extract effective subject key if it follows the format "Access Management Service Request | Actual Subject"
            effective_subject_key = subject.split("|")[-1].strip() if "access management service request |" in subject else subject

            logger.info(f"Subject: {subject} Effective subject key: {effective_subject_key}")
            # key: email subject, value: excel sheet
            subject_map = {
                "new user account creation": "New User Account",
                "new service account creation": "New Service Account",
                "user account modification": "User Account Modify",
                "service account modification": "Service Account Modify",
                "disable service account": "Disable Service Account",
                "disable user account": "Disable User Account",
                "user account password reset": "User Password Reset",
                "service account password reset": "ServiceAc Passwd Reset",
                "other requests": "Other Requests",
            }

            # Matching email subject to sheet name using effective subject key
            matched_key = next((key for key in subject_map if key in effective_subject_key), None)
            if matched_key:
                logger.info(f"Matched key {matched_key}")
                sheet_name = subject_map[matched_key]
                logger.info(f"conmtent {attachment_content} sheet_name {sheet_name} subject {subject}")
                logger.info("Converting the user data excel to json")
                processed_data = user_data_excel_to_json(attachment_content, sheet_name, subject)
            else:

                # evalidate if the email subject conatins Access Management Support Data
                if "access management support" in subject and ENABLE_AM_SUPPORT_DATA_PROCESSING:
                    logger.info("Saving the support excel data into oscar accessmanagement db")
                    worklog_id = worklog_hook.worklog_id
                    processed_data = save_mapping_excel_data_to_db(attachment_content, subject, worklog_id)
                else:
                    logger.info("Converting the user data excel to json")
                    processed_data = mapping_excel_to_json(attachment_content)

            if processed_data and processed_data.get("status") == "success":
                status = processed_data.get("status", "success")
                worklog_hook.info(f"Successfully processed Excel file: {filename}")
                logger.info(f"Successfully processed Excel file: {filename}")
                return {
                    "filename": filename,
                    "content_type": content_type,
                    "status": status.lower(),
                    "processed_data": processed_data,
                }
            else:
                error_message = processed_data.get("message", "Unknown error")
                logger.error(f"Error processing Excel file: {error_message}")
                worklog_hook.error(f"Error processing Excel file: {error_message}")
                return {"filename": filename, "content_type": content_type, "status": "error", "error": error_message}
        except Exception as e:
            worklog_hook.error(f"Error processing Excel file: {str(e)}")
            logger.error(f"Error processing Excel file: {str(e)}")
            return {
                "filename": filename,
                "content_type": content_type,
                "status": "error",
                "error": f"Error processing Excel file: {str(e)}",
            }
    else:
        # For other content types, check if it's text-based
        is_text = any(
            content_type.startswith(prefix)
            for prefix in ["text/", "application/json", "application/xml", "application/csv"]
        )

        if is_text:
            try:
                logger.info("processing text based attachment")
                # Decode the base64 content
                decoded_content = base64.b64decode(attachment_content)
                text_content = decoded_content.decode("utf-8")
                lines = text_content.splitlines()
                total_lines = len(lines)

                worklog_hook.info(f"Total lines in {filename}: {total_lines}")
                logger.info(f"Total lines in {filename}: {total_lines}")
                worklog_hook.info(f"Total size: {size} bytes")
                logger.info(f"Total size: {size} bytes")

                # Log first 5 lines
                worklog_hook.info(f"First 5 lines of {filename}:")
                logger.info(f"First 5 lines of {filename}:")
                for line in lines[:5]:
                    worklog_hook.info(line)
                    logger.info(line)

                # Log last 5 lines if there are more than 10 lines
                if total_lines > 10:
                    worklog_hook.info(f"Last 5 lines of {filename}:")
                    logger.info(f"Last 5 lines of {filename}:")
                    for line in lines[-5:]:
                        worklog_hook.info(line)
                        logger.info(line)
                elif total_lines > 5:
                    worklog_hook.info(f"Remaining lines of {filename}:")
                    logger.info(f"Remaining lines of {filename}:")
                    for line in lines[5:]:
                        worklog_hook.info(line)
                        logger.info(line)

                return {
                    "filename": filename,
                    "content_type": content_type,
                    "status": "ok",
                    "total_lines": total_lines,
                    "size": size,
                }
            except Exception as e:
                worklog_hook.error(f"Error processing text file: {str(e)}")
                logger.error(f"Error processing text file: {str(e)}")
                return {
                    "filename": filename,
                    "content_type": content_type,
                    "status": "error",
                    "error": f"Error processing text file: {str(e)}",
                }
        else:
            # For non-text files, just log the size
            worklog_hook.info(f"Non-text file: {filename}")
            logger.info(f"Non-text file: {filename}")
            worklog_hook.info(f"Content Type: {content_type}")
            logger.info(f"Content Type: {content_type}")
            worklog_hook.info(f"Size: {size} bytes")
            logger.info(f"Size: {size} bytes")
            return {"filename": filename, "content_type": content_type, "status": "ok", "size": size}


# Function to encode the Excel file into base64 string
def encode_file_to_base64(file_path):
    with open(file_path, "rb") as file:
        # Read the file and encode it to base64
        encoded_string = base64.b64encode(file.read()).decode("utf-8")
        logger.info(f"Encoded string: {encoded_string}")
    return encoded_string


# Function to encode strings into base64


def encode_to_base64(input_string):
    if isinstance(input_string, str):
        # Encode the string into bytes, then encode it to base64
        encoded_bytes = base64.b64encode(input_string.encode("utf-8"))
        logger.info(f"Encoded bytes: {encoded_bytes}")
        return encoded_bytes.decode("utf-8")  # Decode back to UTF-8 string
    return input_string


# Function to convert the base64 encoded string back into JSON


def mapping_excel_to_json(base64_string):
    try:
        # Decode the base64 string to bytes
        decoded_bytes = base64.b64decode(base64_string)

        # Create a BytesIO stream from the decoded bytes
        workbook_file = BytesIO(decoded_bytes)

        # Read the Excel file using pandas; this reads all sheets into a dictionary of DataFrames
        workbook_dict = pd.read_excel(workbook_file, sheet_name=None)

        # Construct the JSON structure
        workbook_json = {"workbook": {"sheets": []}}

        for sheet_name, df in workbook_dict.items():
            # Check if it's a Series (one column sheet) and convert it to DataFrame
            if isinstance(df, pd.Series):
                df = df.to_frame()

            # Encode headers and rows into base64
            sheet_data = {
                "name": sheet_name,
                "headers": [encode_to_base64(header) for header in df.columns],
                "rows": [],
            }

            for row in df.to_dict(orient="records"):
                # For each cell in the row, encode the value to base64
                sanitized_row = {key: encode_to_base64(value) for key, value in row.items()}
                sheet_data["rows"].append(sanitized_row)

            workbook_json["workbook"]["sheets"].append(sheet_data)

        # Convert the dictionary to a JSON string
        json_string = json.dumps(workbook_json, indent=4)
        logger.info(f"JSON string after excel to json conversion: {json_string}")
        return json_string
    except Exception as e:
        logger.error(f"Error in mapping_excel_to_json: {str(e)}")
        return {"status": "error", "message": f" An error occurred: {str(e)}"}


def user_data_excel_to_json(base64_string, sheet_name, subject):
    # the following map will have the actual json fields that access management processor qill recieve while getting the request data fro an acccess_management_request processed from the request sheet
    EXCEL_TO_JSON_MAP = {
        "User's full name": "user_full_name",
        "User's Full Name": "user_full_name",
        "User's NTID (for T-Mobile user) / Signum (for Ericsson user)": "user_ntid_signum",
        "User's Email Id": "user_email_id",
        "User's Uprising ID (AD ID)": "user_uprising_id",
        "Official Mobile number (if available)": "official_mobile_no",
        "Service account full name": "service_account_full_name",
        "Service Account Owner Name": "service_account_owner_name",
        "Service Account Owner Email": "service_account_owner_email_id",
        "Service Account Owner Email ID": "service_account_owner_email_id",
        "Service Account Owner Uprising Id (AD ID)": "service_account_owner_uprising_id",
        "Service Account Owner Manager's AD ID": "manager_ad_id",
        "Service Account Owner NTID (for T-Mobile user) / Signum (for Ericsson user)": "owner_ntid",
        "Account Owner group email(if any)": "owner_group_email",
        "Environment (Prod/Lab)": "environment",
        "Owner Group Email (If available)": "owner_group_email",
        "Organization (T-Mobile/Ericsson)": "organization",
        "Owner Official Mobile No (If available)": "official_mobile_no",
        "Official Mobile No (If available)": "official_mobile_no",
        "Reporting Manager name": "manager_name",
        "Reporting Manager email": "manager_email_id",
        "Reporting Manager's email ID": "manager_email_id",
        "Reporting Manager's AD ID": "manager_ad_id",
        "Team name": "team_name",
        "Business Justification (Please clearly specify your job role and purpose of the access request)": "business_justification",
        "Reference AD ID for mirroring of accesses (MUST be from the same Team/Job Role)": "reference_ad_id",
        "Environment ( Please select only if you exactly know the environment, you may confirm it from Reference AD ID of your colleague in ITSM)": "environment",
        "Description/Keyword ( If you don't know exact role/enviroment, then please select from this keyword/description which best suits your purpose)": "description",
        "Description/Keyword ( If you don't know exact role/enviroment, then please select from this keyword/description which best suits your purpose": "description",
        "Description/Keyword ( Please select from this keyword/description which best suits your purpose)": "description",
        "Action (Add/Remove)": "action",
        "Intended Use": "intended_use",
        "Service Request number raised in ITSM": "service_request_number",
        "New Service Owner Full Name (If the user is owner of any service account. To be filled by FO)": "service_account_owner_name",
        "New Service Owner Email ID (If the user is owner of any service account. To be filled by FO)": "service_account_owner_email_id",
        "New Service Owner AD Id (for Prod) (If the user is owner of any service account. To be filled by FO)": "service_owner_ad_id_prod",
        "New Service Owner AD Id (for Lab) (If the user is owner of any service account. To be filled by FO)": "service_owner_ad_id_lab",
        "User's AD ID": "user_ad_id",
        "User Official Mobile No (if Available)": "official_mobile_no",
        "User's Email Address": "user_email_id",
        "Organization (TMobile/Ericsson)": "organization",
        "Environment (Lab/Prod)": "environment",
        "Environment (LAB/PROD)": "environment",
        "Comments": "comments",
        "Comments (If any)": "comments",
        "Platform": "platform",
        "User's Manager AD ID": "manager_ad_id",

    }
    try:
        # Decoding base64string to bytes and loading workbook
        logger.info(f"Decoding base64string to bytes and loading workbook")
        excel_bytes = base64.b64decode(base64_string)
        excel_io = BytesIO(excel_bytes)
        wb = openpyxl.load_workbook(excel_io, data_only=True)

        # validating the sheet
        normalized_sheetnames = [name.lower() for name in wb.sheetnames]
        if sheet_name.lower() not in normalized_sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook.")
            return {"status": "error", "message": f" Sheet '{sheet_name}' not found in workbook."}

        # Get the original sheet name to maintain case sensitivity in the workbook
        original_sheet_name = wb.sheetnames[normalized_sheetnames.index(sheet_name.lower())]
        sheet = wb[original_sheet_name]

        # Extract data
        output_data = {}
        empty_fields = []

        for row in sheet.iter_rows(min_row=3, max_col=2):
            key_cell = row[0].value
            value_cell = row[1].value

            if key_cell:
                key = str(key_cell).strip()
                value = str(value_cell).strip() if value_cell else ""
                output_data[key] = value

                if value == "":
                    empty_fields.append(key)

        # Validating values
        if empty_fields:
            return {
                "status": "error",
                "missing_fields": empty_fields,
                "message": f"The following fields are missing values: {', '.join(empty_fields)}",
            }

        # Map Excel keys to desired JSON keys, but keep original key if not in the map
        output_mapped_data = {
            EXCEL_TO_JSON_MAP[k] if k in EXCEL_TO_JSON_MAP else k: v
            for k, v in output_data.items()
        }
        # Add email_subject to the output_mapped_data
        output_mapped_data["email_subject"] = subject
        logger.info(f"Output mapped data: {output_mapped_data}")
        return {"status": "success", "data": output_mapped_data, "access_form": True}

    except Exception as e:
        return {"status": "error", "message": f" An error occurred: {str(e)}"}


def validate_access_management_data(output_mapped_data: Dict[str, Any], subject: str, db_hook=None) -> Dict[str, Any]:
    """
    Validate access management request data based on specific field requirements.

    Args:
        output_mapped_data: The processed data from the access form
        subject: The original email subject
        db_hook: Database hook for AD ID validation

    Returns:
        Dict containing validation results with status and failed fields
    """
    failed_fields = []
    validation_errors = []
    logger.info(f"Validating access management data for subject: {subject}")
    # Validate each field based on its name and content
    for field_name, field_value in output_mapped_data.items():
        if not field_value:  # Skip empty values
            continue

        field_name_lower = field_name.lower()
        field_value_str = str(field_value).strip()

        # Name fields - max 50 characters
        if 'name' in field_name_lower:
            if len(field_value_str) > 60:
                failed_fields.append(field_name)
                validation_errors.append(f"{field_name}: Value exceeds 30 characters limit")

        # Email fields - basic email validation
        elif 'email' in field_name_lower and 'subject' not in field_name_lower:
            if '@' not in field_value_str or '.' not in field_value_str:
                failed_fields.append(field_name)
                logger.error(f"{field_name}: Invalid email format - {field_value_str}")
                validation_errors.append(f"{field_name}: Invalid email format")

        # NT ID fields - max 10 characters
        elif 'nt_id' in field_name_lower or 'ntid' in field_name_lower:
            if len(field_value_str) > 10:
                failed_fields.append(field_name)
                logger.error(f"{field_name}: Value exceeds 10 characters limit - {field_value_str}")
                validation_errors.append(f"{field_name}: Value exceeds 10 characters limit")

        # AD ID fields - max 10 characters and database validation
        elif 'ad_id' in field_name_lower or 'adid' in field_name_lower:
            if len(field_value_str) > 10:
                failed_fields.append(field_name)
                logger.error(f"{field_name}: Value exceeds 10 characters limit - {field_value_str}")
                validation_errors.append(f"{field_name}: Value exceeds 10 characters limit")
            else:
                # If length validation passes, check if AD ID exists in database
                if db_hook:
                    try:
                        user_name_adid, user_email_adid = db_hook.get_info_by_adid(field_value_str)
                        if not user_name_adid or not user_email_adid:
                            failed_fields.append(field_name)
                            logger.error(f"{field_name}: AD ID '{field_value_str}' not found in database")
                            validation_errors.append(f"{field_name}: AD ID '{field_value_str}' not found in database")
                    except Exception as e:
                        failed_fields.append(field_name)
                        logger.error(f"Error validating {field_name} with value '{field_value_str}' in database: {str(e)}")
                        validation_errors.append(f"{field_name}: Error validating AD ID '{field_value_str}' in database: {str(e)}")

        # Organization fields - must contain specific values
        elif 'organization' in field_name_lower:
            valid_orgs = ['ericsson', 't-mobile']
            if field_value_str.lower() not in valid_orgs:
                failed_fields.append(field_name)
                logger.error(f"{field_name}: Must be one of {', '.join(valid_orgs)}")
                validation_errors.append(f"{field_name}: Must be one of {', '.join(valid_orgs)}")

        # Description fields - check if role and platform can be retrieved from database
        elif 'description' in field_name_lower:
            if db_hook:
                try:
                    # Get organization value for the role_platform lookup
                    organization_value = output_mapped_data.get('organization', '')
                    u_role, u_platform = db_hook.get_role_platform(field_value_str, organization_value)

                    if not u_role or not u_platform:
                        failed_fields.append(field_name)
                        logger.error(f"{field_name}: No role/platform found for description '{field_value_str}' with organization '{organization_value}'")
                        validation_errors.append(
                            f"{field_name}: Description '{field_value_str}' and Organization '{organization_value}' "
                            f"combination has no platform or role present. Please validate either the description or "
                            f"organization, or contact administrator to check platform role mapping for this specific "
                            f"description-organization combination."
                        )
                except Exception as e:
                    failed_fields.append(field_name)
                    logger.error(f"{field_name}: Error validating description '{field_value_str}' in database: {str(e)}")
                    validation_errors.append(f"{field_name}: Error validating description '{field_value_str}' in database: {str(e)}")

        # Subject-based validation for New User Account requests
        if 'new user account' in subject.lower() and 'description' in field_name_lower:
            field_value_lower = field_value_str.lower()
            if not ('itsm support staff' in field_value_lower or 'itsm-spprt default group for all uprising users' in field_value_lower):
                failed_fields.append(field_name)
                logger.error(f"{field_name}: New User Account request must contain 'ITSM Support Staff' or 'ITSM-SPPRT default group for all uprising users'")
                validation_errors.append(f"{field_name}: New User Account request must contain 'ITSM Support Staff' or 'ITSM-SPPRT default group for all uprising users'")

    logger.info(f"Validating access management data for subject: {subject}")
    # Check if validation passed
    validation_passed = len(failed_fields) == 0

    return {
        "validation_passed": validation_passed,
        "failed_fields": failed_fields,
        "validation_errors": validation_errors
    }


async def process_access_management_requests(**context):
    """
    This DAG task processes an access management request by writing request details to a worklog,
    evaluating the request for alert criteria, and sending an alert via AlertHook if the evaluation passes.
    """
    # Retrieve email data from dag_run.conf.
    dag_run = context.get("dag_run")
    conf = dag_run.conf if dag_run and dag_run.conf else {}
    request_data: Dict[str, Any] = conf.get("email_data", {})

    if not request_data:
        logger.error("Missing request_data in DAG run configuration")
        return None

    # Extract email information.
    subject = request_data.get("subject", "(No Subject)")
    from_addr = request_data.get("from", "(No From Address)")
    to_addr = request_data.get("to", "(No To Address)")
    date = request_data.get("date", str(datetime.now(timezone.utc)))
    body = request_data.get("body", "")
    attachments = request_data.get("attachments", [])

    logger.info(f"subject:{subject} from_address {from_addr} to_address {to_addr} date {date} body {body} attachments available: {'Yes' if attachments else 'No'}")

    # Create or re-use worklog.
    worklog_hook = WorkLogHook()
    worklog_id = conf.get("worklog_id")
    if worklog_id:
        worklog_hook.set_worklog_id(worklog_id)
        worklog_hook.info(f"Using existing worklog with ID: {worklog_id}")
        logger.info(f"Using existing worklog with ID: {worklog_id}")
    else:
        initial_worklog = worklog_hook.create_worklog(
            name="Access Management Request Worklog", description="Worklog for processing access management requests"
        )
        worklog_id = initial_worklog["id"]

    # Log basic email info to the worklog.
    worklog_hook.info(f"Received request: {subject}")
    logger.info(f"Received request: {subject}")
    worklog_hook.info(f"From: {from_addr}")
    logger.info(f"From: {from_addr}")
    worklog_hook.info(f"To: {to_addr}")
    logger.info(f"To: {to_addr}")
    worklog_hook.info(f"Date: {date}")
    logger.info(f"Date: {date}")
    if body:
        worklog_hook.info("Request Body:")
        max_chunk_size = 1000
        if len(body) > max_chunk_size:
            chunks = [body[i : i + max_chunk_size] for i in range(0, len(body), max_chunk_size)]
            for i, chunk in enumerate(chunks):
                worklog_hook.info(f"Body Part {i+1}/{len(chunks)}: {chunk}")
                logger.info(f"Body Part {i+1}/{len(chunks)}: {chunk}")
        else:
            worklog_hook.info(body)
            logger.info(body)
    else:
        worklog_hook.info("No request body content")
        logger.info("No request body content")

    # Process attachments
    processed_attachments = []
    if attachments:
        worklog_hook.info(f"Processing {len(attachments)} attachments")
        logger.info(f"Processing {len(attachments)} attachments")

        for attachment in attachments:
            logger.info(f"attachment to be processed: {attachment}")
            processed_attachment = await process_attachment(attachment, request_data, worklog_hook)
            logger.info("attachment processed")
            processed_attachments.append(processed_attachment)

        # Notify requester if any attachment failed to be retrieved
        failed_attachments = [a for a in processed_attachments if a.get("status") == "error"]
        if failed_attachments:
            failed_filename = failed_attachments[0].get("filename", "attachment")
            failed_error = failed_attachments[0].get("error", "Unknown error")
            logger.error(f"Attachment {failed_filename!r} failed: {failed_error!r} — notifying {from_addr}")
            worklog_hook.error(f"Attachment {failed_filename} could not be retrieved. Notifying requester.")
            try:
                # Try to load a custom template from mapping; fall back to hardcoded if unavailable
                template_content = ""
                try:
                    mapping_hook = MappingHook()
                    mapping_elements = mapping_hook.list_mapping_elements(
                        mapping_name="access_management_sr_management",
                        mapping_namespace_name="access_management_enable",
                        mapping_key="email_request_attachment_error_email_template"
                    )
                    if mapping_elements and isinstance(mapping_elements[0], dict):
                        template_content = mapping_elements[0].get("value", "")
                except Exception as tmpl_err:
                    logger.warning(f"Could not load attachment error email template from mapping: {tmpl_err}. Using fallback.")

                if not template_content:
                    template_content = (
                        "Dear Requester,\n\n"
                        "We received your access management request but were unable to process it due to an attachment retrieval error.\n\n"
                        "Details:\n"
                        "  From       : {{ from_addr }}\n"
                        "  Subject    : {{ subject }}\n"
                        "  Received   : {{ date }}\n"
                        "  Attachment : {{ filename }}\n\n"
                        "This is usually a transient system error. Please resend your email with the attachment and it will be processed automatically.\n\n"
                        "If the issue persists, please contact the OSCAR support team.\n\n"
                        "Regards,\nOSCAR Access Management"
                    )

                rendered_content = Template(template_content).render(
                    subject=subject,
                    from_addr=from_addr,
                    to_addr=to_addr,
                    date=date,
                    filename=failed_filename,
                    error_detail=failed_error,
                )
                notify_hook = NotifyHook()
                notifier_name = os.environ.get("OSCAR_ENABLE_NOTFIER_NAME", "oscar_notifier_email")
                email_group = os.environ.get("OSCAR_ENABLE_EMAIL_GROUP", "am_email_group")
                result = notify_hook.send_notification({
                    "name": notifier_name,
                    "subject": f"{subject} | Processing Failed — Please Resend",
                    "message": rendered_content,
                    "cc_notifier_id": email_group,
                    "notifier_id": [from_addr],
                })
                logger.info(f"Attachment error notification sent to {from_addr}: {result}")
            except Exception as notify_err:
                logger.error(f"Failed to send attachment error notification: {notify_err}")
            worklog_hook.close_worklog()
            return {
                "worklog_id": worklog_id,
                "status": "failed",
                "email_subject": subject,
                "reason": f"Attachment {failed_filename!r} could not be retrieved: {failed_error}",
            }
    else:
        worklog_hook.info("No attachments to process")
        if "access management" in subject.lower() and "|" in subject.lower():
            logger.debug("access management request with no attachment is received")

            try:
                # Initialize mapping hook
                mapping_hook = MappingHook()

                # Fetch template using mapping hook
                mapping_elements = mapping_hook.list_mapping_elements(
                    mapping_name="access_management_sr_management",
                    mapping_namespace_name="access_management_enable",
                    mapping_key="email_request_attachment_missing_email_template"
                )

                if not mapping_elements:
                    logger.error("No mapping element for email template found for attachment less access management request")
                    raise ValueError("No mapping element for email template found for attachment less access management request")

                # Get the first matching element and extract its value
                template_element = mapping_elements[0]
                if not isinstance(template_element, dict):
                    logger.error("Invalid mapping element format")
                    raise ValueError("Invalid mapping element format")

                template_content = template_element.get("value")
                if not template_content:
                    logger.error("No template content found in mapping element")
                    raise ValueError("No template content found in mapping element")

                # Transform data to match template placeholders
                template_data = {
                    "subject": subject,
                    "from_addr": from_addr,
                    "to_addr": to_addr,
                    "date": date
                }

                # Render content
                rendered_content = Template(template_content).render(**template_data)            

                notify_hook = NotifyHook()
                notifier_name = os.environ.get("OSCAR_ENABLE_NOTFIER_NAME", "oscar_notifier_email")
                email_group = os.environ.get("OSCAR_ENABLE_EMAIL_GROUP", "am_email_group")
                user_email = from_addr
                result = notify_hook.send_notification({
                    "name": notifier_name,
                    "subject": subject,
                    "message": rendered_content,
                    "cc_notifier_id": email_group,  # fo_email should be in email group
                    "notifier_id": [user_email]
                })

                logger.info(f"Notification task async initiated with initial result: {result}")
                logger.info(f"Notification task initiated for attachment less access management request")

            except Exception as e:
                logger.error(f"Error sending access management notification: {str(e)}")

        logger.info(f"No attachments found to process for request {request_data}")

    # Finish up worklog logging.
    worklog_hook.info("Finished writing request details to worklog.")

    closed_worklog = worklog_hook.close_worklog()
    # Instead of using worklog_hook.info (which fails since the worklog is closed),
    # use the standard logger to log the closed worklog ID.
    logger.info(f"Worklog closed with ID: {closed_worklog['id']}")

    # Extract data from first processed attachment if it exists

    attachment_count = len(processed_attachments) if processed_attachments is not None else 0
    attachment_status = "None" if processed_attachments is None else str(processed_attachments)
    logger.info(f"Proceeding to call access_management_sr_management DAG based on processed_attachments: {attachment_count} - {attachment_status}")

    # Start a saperate worklog as the main worklog is closed
    new_worklog_hook = WorkLogHook()
    new_worklog = new_worklog_hook.create_worklog(
        name="Access Management Request Worklog", description="Worklog for processing access management requests"
    )
    new_worklog_id = new_worklog["id"]

    if processed_attachments:
        logger.info("Processed attachment found")
        # Filter for spreadsheet attachments only
        processed_spreadsheet_attachments = [
            attachment for attachment in processed_attachments
            if attachment.get("content_type") in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]
            or any(attachment.get("content_type", "").startswith(prefix) for prefix in ["text/", "application/json", "application/xml", "application/csv"])
        ]
        logger.info(f"only one attachment is supported as email subject line reflects the type of request- proceeding to process the first attachment")
        processed_attachment = processed_spreadsheet_attachments[0]
        processed_data = processed_attachment.get("processed_data", {})
        logger.debug(f"the processed data from attachment {processed_data}")

        # Check if this is an access form with successful processing
        logger.info(f"processed data : {processed_data}")
        if (processed_data.get("access_form") and processed_data.get("status") == "success" and "data" in processed_data):

            logger.info("access form data found and proceeding to call access_management_sr_management DAG")
            output_mapped_data = processed_data["data"]
            logger.info(f"{closed_worklog['id']} Successfully extracted output_mapped_data from access form: {output_mapped_data}")
            logger.info("Found valid access form data for further processing")

            # Validate the data before triggering the DAG
            oscr_connection_id = os.environ.get("OSCAR_DB_EXT_CONNECTION_ID", "oscar_db_ext")
            db_hook = AccessManagementSQLHook(connection_id=oscr_connection_id, worklog_id=new_worklog_id)
            validation_result = validate_access_management_data(output_mapped_data, subject, db_hook)

            if validation_result["validation_passed"]:
                logger.info("Data validation passed, proceeding to trigger access_management_sr_management DAG")

                # Trigger access_management_sr_management DAG via middleware API
                try:
                    # Generate a unique DAG run ID using timestamp
                    current_time = datetime.now(timezone.utc).isoformat()
                    dag_run_id = f"access_management_sr_{current_time}"

                    # Prepare payload for the workflow
                    payload = {
                        "dag_run_id": dag_run_id,
                        "conf": output_mapped_data,
                        "note": f"Access management SR request triggered from process_access_management_requests"
                    }

                    # Get middleware API details from environment/config
                    MIDDLEWARE_HOST: str = os.environ.get("MIDDLEWARE_HOST", "middleware")
                    MIDDLEWARE_PORT: int = int(os.environ.get("MIDDLEWARE_PORT", 5200))
                    access_management_sr_creation_workflow_id = os.environ.get("ACCESS_MANAGEMENT_SR_MANAGEMENT", "access_management_sr_management")

                    middleware_url = f"https://{MIDDLEWARE_HOST}:{MIDDLEWARE_PORT}"

                    logger.info(f"middleware url for airflow dag: {middleware_url}/api/v1/workflows/{access_management_sr_creation_workflow_id}")
                    logger.info(f"middleware payload: {payload}")

                    async def trigger_workflow():
                        logger.info(f"triggering access_management_sr_management DAG with payload: {payload}")
                        async with httpx.AsyncClient(verify=False) as client:
                            response = await client.post(
                                f"{middleware_url}/api/v1/workflows/{access_management_sr_creation_workflow_id}",
                                json=payload
                            )
                            response.raise_for_status()
                            return response.json()

                    result = await trigger_workflow()
                    logger.info(f"Successfully triggered access_management_sr_management DAG with run ID: {result.get('dag_run_id')}")

                except Exception as e:
                    logger.error(f"Error triggering access_management_sr_management DAG: {str(e)}")
            else:
                logger.warning(f"Data validation failed: {validation_result['validation_errors']}")
                # Send notification about validation failures
                try:
                    # Initialize mapping hook
                    mapping_hook = MappingHook()

                    # Fetch template using mapping hook
                    mapping_elements = mapping_hook.list_mapping_elements(
                        mapping_name="access_management_sr_management",
                        mapping_namespace_name="access_management_enable",
                        mapping_key="email_request_validation_failure_email_template"
                    )

                    if not mapping_elements:
                        logger.error("No mapping element for validation failure email template found")
                        raise ValueError("No mapping element for validation failure email template found")

                    # Get the first matching element and extract its value
                    template_element = mapping_elements[0]
                    if not isinstance(template_element, dict):
                        logger.error("Invalid mapping element format")
                        raise ValueError("Invalid mapping element format")

                    template_content = template_element.get("value")
                    if not template_content:
                        logger.error("No template content found in mapping element")
                        raise ValueError("No template content found in mapping element")

                    # Transform data to match template placeholders
                    def get_friendly_field_name(field):
                        return FIELD_NAME_MAP.get(field, field)
                    template_data = {
                        "subject": subject,
                        "from_addr": from_addr,
                        "failed_fields": ", ".join(get_friendly_field_name(f) for f in validation_result["failed_fields"]),
                        "validation_errors": " | ".join(validation_result["validation_errors"]),
                        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
                    }

                    # Render content
                    rendered_content = Template(template_content).render(**template_data)

                    notify_hook = NotifyHook()
                    notifier_name = os.environ.get("OSCAR_ENABLE_NOTFIER_NAME", "oscar_notifier_email")
                    email_group = os.environ.get("OSCAR_ENABLE_EMAIL_GROUP", "am_email_group")
                    user_email = from_addr
                    result = notify_hook.send_notification({
                        "name": notifier_name,
                        "subject": f"{subject} | Validation Failed",
                        "message": rendered_content,
                        "cc_notifier_id": email_group,
                        "notifier_id": [user_email]
                    })

                    logger.info(f"Validation failure notification sent: {result}")
                    logger.info(f"Validation failure notification sent for failed fields: {validation_result['failed_fields']}")

                except Exception as e:
                    logger.error(f"Error sending validation failure notification: {str(e)}")
                    logger.error(f"Error sending validation failure notification: {str(e)}")

            # Close the new worklog
            new_worklog = new_worklog_hook.close_worklog()
            logger.info(f"New worklog closed with ID: {new_worklog['id']}")

        elif (not processed_data.get("access_form") and
              processed_data.get("status") == "success" and
              processed_data.get("excel_type")):
            # Support table insert succeeded (platform_role_mapping / shift_roster / user_report / approver_role)
            # The access_form path above handles its own notification; this path was silently returning with none.
            excel_type = processed_data.get("excel_type", "")
            rows_processed = processed_data.get("rows_processed", 0)
            table_name = processed_data.get("table", excel_type)
            friendly_type_map = {
                "platform_role_mapping": "Platform Role Mapping",
                "approver_role": "Approver Role",
                "user_report": "User Report (ITSM User List)",
                "shift_roster": "Shift Roster",
            }
            friendly_type = friendly_type_map.get(excel_type, excel_type.replace("_", " ").title())
            logger.info(f"Support data insert succeeded: type={excel_type}, rows={rows_processed} — notifying {from_addr}")
            new_worklog_hook.info(f"Support data '{friendly_type}' successfully inserted ({rows_processed} rows). Notifying requester.")
            try:
                template_content = ""
                try:
                    mapping_hook = MappingHook()
                    mapping_elements = mapping_hook.list_mapping_elements(
                        mapping_name="access_management_sr_management",
                        mapping_namespace_name="access_management_enable",
                        mapping_key="email_request_support_data_success_email_template"
                    )
                    if mapping_elements and isinstance(mapping_elements[0], dict):
                        template_content = mapping_elements[0].get("value", "")
                except Exception as tmpl_err:
                    logger.warning(f"Could not load support data success email template from mapping: {tmpl_err}. Using fallback.")

                if not template_content:
                    template_content = (
                        "<p>Dear Requester,</p>"
                        "<p>Your access management support data has been successfully processed and updated in the system.</p>"
                        "<p><strong>Details:</strong></p>"
                        "<table style='border-collapse:collapse;'>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>From</strong></td><td>{{ from_addr }}</td></tr>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>Subject</strong></td><td>{{ subject }}</td></tr>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>Received</strong></td><td>{{ date }}</td></tr>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>Data Type</strong></td><td>{{ friendly_type }}</td></tr>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>Rows Updated</strong></td><td>{{ rows_processed }}</td></tr>"
                        "<tr><td style='padding:4px 12px 4px 0;'><strong>Table</strong></td><td>{{ table_name }}</td></tr>"
                        "</table>"
                        "<p>No further action is required.</p>"
                        "<p>Regards,<br>OSCAR Access Management</p>"
                    )

                rendered_content = Template(template_content).render(
                    subject=subject,
                    from_addr=from_addr,
                    to_addr=to_addr,
                    date=date,
                    friendly_type=friendly_type,
                    rows_processed=rows_processed,
                    table_name=table_name,
                )
                notify_hook = NotifyHook()
                notifier_name = os.environ.get("OSCAR_ENABLE_NOTFIER_NAME", "oscar_notifier_email")
                email_group = os.environ.get("OSCAR_ENABLE_EMAIL_GROUP", "am_email_group")
                result = notify_hook.send_notification({
                    "name": notifier_name,
                    "subject": f"{subject} | Data Updated Successfully",
                    "message": rendered_content,
                    "notifier_id": [from_addr],
                })
                logger.info(f"Support data success notification sent to {from_addr}: {result}")
            except Exception as notify_err:
                logger.error(f"Failed to send support data success notification: {notify_err}")

            new_worklog = new_worklog_hook.close_worklog()
            logger.info(f"New worklog closed with ID: {new_worklog['id']}")

    return {
        "worklog_id": closed_worklog["id"],
        "status": "processed",
        "email_subject": subject,
        "processed_attachments": processed_attachments,
    }


with DAG(
    dag_id="process_access_management_requests",
    default_args=default_args,
    description="Process access management requests: new user, password reset, etc.",
    schedule=None,  # This DAG is triggered via an API call or upstream process.
    start_date=pendulum.today('UTC').add(days=-1),
    tags=["access", "management", "requests", "processing"],
) as dag:

    process_access_management_requests_task = PythonOperator(
        task_id="process_access_management_requests",
        python_callable=lambda **context: asyncio.run(process_access_management_requests(**context)),
    )
